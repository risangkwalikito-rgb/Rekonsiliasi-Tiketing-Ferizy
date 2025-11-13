# path: streamlit_app.py
import io
import zipfile
from datetime import date
from collections import defaultdict, OrderedDict
from typing import Dict, List, Optional, Tuple

import pandas as pd
import streamlit as st
from openpyxl import load_workbook  # streaming .xlsx read_only


# =========================== Konfigurasi & Konstanta ===========================

COL_H = "TIPE PEMBAYARAN"                      # H
COL_B = "TANGGAL PEMBAYARAN"                   # B
COL_AA = "REF NO"                              # AA
COL_K = "TOTAL TARIF TANPA BIAYA ADMIN (Rp.)"  # K
COL_X = "SOF ID"                               # X
COL_ASAL = "ASAL"                              # (Pelabuhan)
REQUIRED_COLS = [COL_H, COL_B, COL_AA, COL_K, COL_X, COL_ASAL]

CAT_COLS = [
    "Cash",
    "Prepaid BRI",
    "Prepaid BNI",
    "Prepaid Mandiri",
    "Prepaid BCA",
    "SKPT",
    "IFCS",
    "Reedem",
    "ESPAY",
    "Finnet",
]

NON_COMPONENTS = [
    "Cash",
    "Prepaid BRI",
    "Prepaid BNI",
    "Prepaid Mandiri",
    "Prepaid BCA",
    "SKPT",
    "IFCS",
    "Reedem",
]

CSV_CHUNK_ROWS = 200_000
XLSX_BATCH_ROWS = 50_000
VALID_EXTS = (".xlsx", ".xls", ".csv")


# =========================== Utilitas umum ===========================

def _ensure_required_columns(df: pd.DataFrame) -> None:
    missing = [c for c in REQUIRED_COLS if c not in df.columns]
    if missing:
        raise ValueError("Kolom wajib tidak ditemukan: " + ", ".join(missing) + ".")

def _style_table(df_display: pd.DataFrame, highlight: bool) -> "pd.io.formats.style.Styler":
    numeric_cols = df_display.select_dtypes(include="number").columns.tolist()
    styler = df_display.style.format("{:,.0f}", subset=numeric_cols)
    if highlight and "Selisih" in df_display.columns:
        styler = styler.apply(
            lambda s: [
                "background-color:#fdecea; color:#b71c1c; font-weight:600;" if (pd.notna(v) and float(v) != 0) else ""
                for v in s
            ],
            subset=["Selisih"],
        )
    return styler

def _add_subtotal_row(df_display: pd.DataFrame, label: str = "Subtotal", date_col: str = "Tanggal") -> pd.DataFrame:
    numeric_cols = df_display.select_dtypes(include="number").columns.tolist()
    totals = df_display[numeric_cols].sum()
    subtotal = {c: (totals[c] if c in totals else None) for c in df_display.columns}
    subtotal[date_col] = label
    return pd.concat([df_display, pd.DataFrame([subtotal])], ignore_index=True)


# =========================== Agregator streaming (per Tanggal & Pelabuhan) ===========================

def _empty_agg():
    # key: (date, asal) -> {col -> sum}
    return defaultdict(lambda: defaultdict(float))

def _update_agg_series(agg, ser: pd.Series, colname: str) -> None:
    if ser.empty:
        return
    # ser.index: MultiIndex (Tanggal, Asal)
    for (dt, asal), val in ser.items():
        agg[(dt, asal)][colname] += float(val)

def _apply_rules_and_update(df_chunk: pd.DataFrame, agg) -> None:
    # normalisasi
    H = df_chunk[COL_H].fillna("").astype(str).str.lower()
    AA = df_chunk[COL_AA].fillna("").astype(str).str.lower()
    X = df_chunk[COL_X].fillna("").astype(str).str.lower()

    # tampilkan nama pelabuhan tertrim (jangan lower agar terbaca)
    ASAL = df_chunk[COL_ASAL].fillna("Tidak diketahui").astype(str).str.strip()

    amt = pd.to_numeric(df_chunk[COL_K], errors="coerce").fillna(0)
    tgl = df_chunk["Tanggal"]

    def sum_by_key(mask) -> pd.Series:
        if mask.any():
            return amt[mask].groupby([tgl[mask], ASAL[mask]], dropna=False).sum(min_count=1)
        # return MultiIndex-like kosong
        mi = pd.MultiIndex.from_arrays([[], []], names=["Tanggal", "Pelabuhan"])
        return pd.Series(index=mi, dtype="float64")

    # === Kategori utama ===
    rules = OrderedDict([
        ("Cash", H.str.contains("cash", na=False)),
        ("Prepaid BRI", H.str.contains("prepaid-bri", na=False)),
        ("Prepaid BNI", H.str.contains("prepaid-bni", na=False)),
        ("Prepaid Mandiri", H.str.contains("prepaid-mandiri", na=False)),
        ("Prepaid BCA", H.str.contains("prepaid-bca", na=False)),
        ("SKPT", H.str.contains("skpt", na=False)),
        ("IFCS", H.str.contains("ifcs", na=False)),
        ("Reedem", H.str.contains("reedem", na=False) | H.str.contains("redeem", na=False)),
        ("ESPAY", H.str.contains("finpay", na=False) & AA.str.startswith("esp", na=False)),
        ("Finnet", H.str.contains("finpay", na=False) & (~AA.str.startswith("esp", na=False))),
    ])
    for name, m in rules.items():
        _update_agg_series(agg, sum_by_key(m), name)

    # === BCA / NON BCA (finpay + SOF ID) ===
    is_finpay = H.str.contains("finpay", na=False)
    is_bca_tag = X.str.contains("vabcaespay", na=False) | X.str.contains("bluespay", na=False)
    _update_agg_series(agg, sum_by_key(is_finpay & is_bca_tag), "BCA")
    _update_agg_series(agg, sum_by_key(is_finpay & (~is_bca_tag)), "NON BCA")


# =========================== Pembaca cepat (CSV & Excel) ===========================

def _process_csv_fast(data: bytes, year: int, month: int, agg) -> None:
    itr = pd.read_csv(
        io.BytesIO(data),
        usecols=REQUIRED_COLS,
        chunksize=CSV_CHUNK_ROWS,
        dtype={COL_H: "string", COL_AA: "string", COL_X: "string", COL_ASAL: "string"},
    )
    for chunk in itr:
        t = pd.to_datetime(chunk[COL_B], errors="coerce")
        mask = (t.dt.year == year) & (t.dt.month == month)
        if not mask.any():
            continue
        sub = chunk.loc[mask].copy()
        sub["Tanggal"] = t.loc[mask].dt.date
        _apply_rules_and_update(sub, agg)

def _process_xlsx_streaming(data: bytes, year: int, month: int, agg) -> None:
    """Streaming .xlsx (read_only). Fallback ke pandas.read_excel jika gagal (termasuk .xls)."""
    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if header is None:
            wb.close()
            return
        name_to_idx = {str(h).strip(): i for i, h in enumerate(header) if h is not None}
        if not all(c in name_to_idx for c in REQUIRED_COLS):
            wb.close()
            return

        buf = []
        for r in rows:
            try:
                buf.append([
                    r[name_to_idx[COL_H]],
                    r[name_to_idx[COL_B]],
                    r[name_to_idx[COL_AA]],
                    r[name_to_idx[COL_K]],
                    r[name_to_idx[COL_X]],
                    r[name_to_idx[COL_ASAL]],
                ])
            except Exception:
                continue
            if len(buf) >= XLSX_BATCH_ROWS:
                _flush_xlsx_batch(buf, year, month, agg)
                buf.clear()
        if buf:
            _flush_xlsx_batch(buf, year, month, agg)
            buf.clear()
        wb.close()
    except Exception:
        # Fallback: baca full sheet (lebih lambat & boros RAM)
        try:
            df = pd.read_excel(io.BytesIO(data), sheet_name=0, usecols=REQUIRED_COLS)
        except Exception:
            return
        t = pd.to_datetime(df[COL_B], errors="coerce")
        mask = (t.dt.year == year) & (t.dt.month == month)
        if not mask.any():
            return
        sub = df.loc[mask].copy()
        sub["Tanggal"] = t.loc[mask].dt.date
        _apply_rules_and_update(sub, agg)

def _flush_xlsx_batch(buf: List[List], year: int, month: int, agg) -> None:
    df = pd.DataFrame(buf, columns=[COL_H, COL_B, COL_AA, COL_K, COL_X, COL_ASAL])
    t = pd.to_datetime(df[COL_B], errors="coerce")
    mask = (t.dt.year == year) & (t.dt.month == month)
    if not mask.any():
        return
    sub = df.loc[mask].copy()
    sub["Tanggal"] = t.loc[mask].dt.date
    _apply_rules_and_update(sub, agg)


# =========================== Loader multi-file & ZIP (streaming) ===========================

def _load_and_aggregate(files: List["st.runtime.uploaded_file_manager.UploadedFile"], year: int, month: int):
    agg = _empty_agg()
    for f in files:
        try:
            data = f.getvalue()
        except Exception:
            data = f.read()
        name = f.name.lower()

        try:
            if name.endswith(".zip"):
                with zipfile.ZipFile(io.BytesIO(data)) as zf:
                    for m in zf.infolist():
                        if m.is_dir():
                            continue
                        low = m.filename.lower()
                        if not low.endswith(VALID_EXTS):
                            continue
                        content = zf.read(m)
                        if low.endswith((".xlsx", ".xls")):
                            _process_xlsx_streaming(content, year, month, agg)
                        else:
                            _process_csv_fast(content, year, month, agg)
            elif name.endswith((".xlsx", ".xls")):
                _process_xlsx_streaming(data, year, month, agg)
            elif name.endswith(".csv"):
                _process_csv_fast(data, year, month, agg)
        except Exception:
            # skip file yang gagal
            continue
    return agg


# =========================== Build hasil dari aggregator ===========================

def _build_result_from_agg(agg) -> pd.DataFrame:
    if not agg:
        return pd.DataFrame()

    # rows per (Tanggal, Pelabuhan)
    rows: List[dict] = []
    for (dt, asal), bucket in agg.items():
        row = {"Tanggal": dt, "Pelabuhan": asal}
        for c in CAT_COLS:
            row[c] = bucket.get(c, 0.0)
        row["Total"] = sum(row[c] for c in CAT_COLS)
        bca = bucket.get("BCA", 0.0)
        nonbca = bucket.get("NON BCA", 0.0)
        row["BCA"] = bca
        row["NON BCA"] = nonbca
        row["NON"] = sum(row[c] for c in NON_COMPONENTS)
        row["TOTAL"] = bca + nonbca + row["NON"]
        row["Selisih"] = row["TOTAL"] - row["Total"]
        rows.append(row)

    df = pd.DataFrame(rows)
    if df.empty:
        return df
    # urut Pelabuhan, Tanggal
    df = df[["Tanggal", "Pelabuhan"] + CAT_COLS + ["Total", "BCA", "NON BCA", "NON", "TOTAL", "Selisih"]]
    df = df.sort_values(["Pelabuhan", "Tanggal"]).reset_index(drop=True)
    return df


# =========================== Streamlit UI ===========================

def _to_excel_bytes(df: pd.DataFrame, sheet_name: str = "Rekonsiliasi") -> Tuple[Optional[bytes], Optional[str], Optional[str]]:
    for engine in ("xlsxwriter", "openpyxl"):
        try:
            buf = io.BytesIO()
            with pd.ExcelWriter(buf, engine=engine) as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
            return buf.getvalue(), engine, None
        except ImportError:
            continue
        except Exception as e:
            return None, None, f"Gagal menulis Excel dengan {engine}: {e}"
    return None, None, "Tidak ada engine Excel (xlsxwriter/openpyxl). Tambahkan ke requirements."


def _render_port_table(port_name: str, df_port: pd.DataFrame, highlight: bool) -> None:
    # format tanggal & subtotal & angka bulat
    df_show = df_port.copy()
    df_show["Tanggal"] = pd.to_datetime(df_show["Tanggal"]).dt.strftime("%d/%m/%Y")
    df_show = _add_subtotal_row(df_show, label="Subtotal", date_col="Tanggal")
    numeric_cols = df_show.select_dtypes(include="number").columns
    df_show[numeric_cols] = df_show[numeric_cols].round(0).astype("Int64")
    try:
        st.dataframe(_style_table(df_show, highlight=highlight), use_container_width=True)
    except Exception:
        st.dataframe(df_show, use_container_width=True)


def main() -> None:
    st.set_page_config(page_title="Rekonsiliasi Payment Report", layout="wide")
    st.title("Rekonsiliasi Payment Report")

    # Sidebar: filter & uploader
    today = date.today()
    years_options = list(range(today.year - 5, today.year + 6))
    year = st.sidebar.selectbox("Tahun", options=years_options, index=years_options.index(today.year))
    month_names = {
        1: "01 - Januari", 2: "02 - Februari", 3: "03 - Maret", 4: "04 - April",
        5: "05 - Mei", 6: "06 - Juni", 7: "07 - Juli", 8: "08 - Agustus",
        9: "09 - September", 10: "10 - Oktober", 11: "11 - November", 12: "12 - Desember",
    }
    month = st.sidebar.selectbox("Bulan", options=list(range(1, 13)), index=today.month - 1, format_func=lambda m: month_names[m])

    up_files = st.sidebar.file_uploader(
        "Upload ZIP / beberapa Excel (.xlsx/.xls) / CSV",
        type=["zip", "xlsx", "xls", "csv"],
        accept_multiple_files=True,
    )
    highlight = st.sidebar.checkbox("Highlight kolom Selisih ≠ 0", value=True)

    if not up_files:
        st.info("Silakan upload file di panel kiri (bisa banyak file atau ZIP).")
        return

    # Proses streaming semua file (RAM-efisien)
    with st.spinner("Memproses file besar secara streaming…"):
        agg = _load_and_aggregate(up_files, year=year, month=month)

    result = _build_result_from_agg(agg)
    if result.empty:
        st.warning("Tidak ada data valid setelah filter periode & kolom wajib.")
        return

    st.subheader(f"Hasil Rekonsiliasi • Periode: {month_names[month]} {year}")

    # === Split per Pelabuhan (tabs) ===
    ports = list(result["Pelabuhan"].dropna().unique())
    ports.sort()
    tabs = st.tabs(ports if ports else ["(Tidak ada Pelabuhan)"])
    for tab, port in zip(tabs, ports):
        with tab:
            st.markdown(f"**Pelabuhan: {port}**")
            _render_port_table(port, result[result["Pelabuhan"] == port], highlight=highlight)

    # === Unduh gabungan (semua pelabuhan) ===
    st.divider()
    st.subheader("Unduh Hasil (Gabungan Semua Pelabuhan)")

    export_df = result.copy()
    export_df["Tanggal"] = pd.to_datetime(export_df["Tanggal"]).dt.strftime("%d/%m/%Y")
    num_cols = export_df.select_dtypes(include="number").columns
    export_df[num_cols] = export_df[num_cols].round(0).astype("Int64")

    csv_bytes = export_df.to_csv(index=False).encode("utf-8-sig")
    st.download_button("Unduh CSV (Gabungan)", data=csv_bytes, file_name=f"rekonsiliasi_payment_{year}_{month:02d}_per_pelabuhan.csv", mime="text/csv")

    excel_bytes, engine_used, err_msg = _to_excel_bytes(export_df, sheet_name="Rekonsiliasi")
    if excel_bytes:
        st.download_button(
            f"Unduh Excel (.xlsx) (Gabungan){' • ' + engine_used if engine_used else ''}",
            data=excel_bytes,
            file_name=f"rekonsiliasi_payment_{year}_{month:02d}_per_pelabuhan.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    else:
        st.warning("Ekspor Excel dinonaktifkan. Tambahkan `xlsxwriter>=3.1` atau `openpyxl>=3.1` di requirements." + (f"\nDetail: {err_msg}" if err_msg else ""))

    with st.expander("Aturan, Kolom Wajib & Per-Pelabuhan"):
        st.markdown(
            f"""
**Kolom Wajib:** H=**{COL_H}**, B=**{COL_B}**, AA=**{COL_AA}**, K=**{COL_K}**, X=**{COL_X}**, ASAL=**{COL_ASAL}**.

**Split per Pelabuhan:** Tabel dipecah berdasarkan kolom **ASAL**.  
Semua kolom hasil tetap sama: kategori (Cash…Finnet), **Total**, **BCA**, **NON BCA**, **NON**, **TOTAL**, **Selisih** (highlight ≠ 0).  
Subtotal ditampilkan di bawah tiap tabel pelabuhan.
"""
        )


if __name__ == "__main__":
    main()
